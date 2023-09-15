import openpyxl
import sys
wb = openpyxl.load_workbook(".\Squadre.xlsx")
ws = wb.active
s = []; sq = []
offset = 4;n = 0; bil=0
payload = ""
#--------------------------------------------------------------
#Definizione della classe Squadra con i vari dati attribuiti
class Squadra:
    def __init__(self, nome, pg, v, n, p, g, pt):
        self.nome = nome
        self.pg = pg
        self.v = v
        self.n = n
        self.p = p
        self.g = g
        self.pt = pt
    def stampa(self):
        l = 11 - len(self.nome)
        out = self.nome; h = 0
        while h<l:
            out = out + " "
            h=h+1
        print (
        out + " " +
        str(self.pg) + " " +
        str(self.v) + " " +
        str(self.n) + " " +
        str(self.p) + " " +
        self.g + " " +
        str(self.pt) + "\n"
        )
    def gol(self):
        gf = ""
        i=0
        while i<len(self.g) and self.g[i] != ":":
            gf = gf + self.g[i]
            i=i+1
        gs = ""
        i=i+1
        while i<len(self.g):
            gs = gs + self.g[i]
            i=i+1
        res = int(gf) - int(gs)
        return res
#--------------------------------------------------------------
#Definizione funzione compara()

#--------------------------------------------------------------
#Main:
i = 0
try:
    s.append(input("Inserire squadra 1 della serie A:\n"))
except EOFError as e:
    print("Nessun valore inserito\n")
    sys.exit()
try:
    s.append(input("Inserire squadra 2 della serie A:\n"))
except EOFError as e:
    print("Nessun valore inserito\n")
    sys.exit()
print("\n")
i=offset
n = 0
while i < 21 and n<2:
    i = i + 1
    if(str(s[n]) in str(ws.cell(i,2).value)):
        n=n+1
        sq.append(Squadra(str(ws.cell(i,2).value),
        ws.cell(i,3).value,
        ws.cell(i,4).value,
        ws.cell(i,5).value,
        ws.cell(i,6).value,
        str(ws.cell(i,7).value),
        ws.cell(i,8).value))
    if(i==21 and n<2):
        i=0
'''
for el in sq:
    print(el)'''
print("\n\n")


#Confronto dei punti:
if sq[0].pt>sq[1].pt: bil=bil-1
if sq[0].pt<sq[1].pt: bil=bil+1
#Confronto dei gol
if sq[0].gol()>sq[1].gol(): bil=bil-1
if sq[0].gol()<sq[1].gol(): bil=bil+1
print(bil)
print("Dati aggiornati al " + str(ws.cell(1,10).value))
